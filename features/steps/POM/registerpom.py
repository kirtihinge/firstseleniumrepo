import openpyxl

surname=''
mobile=''
password=''

path="C:\KIRTI\MANUAL testing\write da.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)
for i in range (1,rows):
    if sheet.cell(row=i+1,column=2).value=="shalu":
        surname=sheet.cell(row=i+1,column=3).value
        mobile=sheet.cell(row=i+1,column=4).value
        password=sheet.cell(row=i+1,column=5).value


