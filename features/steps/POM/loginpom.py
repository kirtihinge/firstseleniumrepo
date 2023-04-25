import openpyxl


username=''
password=''

path="C:\KIRTI\MANUAL testing\write da.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)
for i in range(1,rows):
    if sheet.cell(row=i+1,column=2).value=="Manager":
    # if sheet.cell(row=i + 1, column=2).value == "fivefinger":
        username=sheet.cell(row=i+1,column=3).value
        password=sheet.cell(row=i+1,column=4).value
        print(username,':---->',password)


