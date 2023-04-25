import time
import openpyxl
from openpyxl.styles import PatternFill, Font
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By

from behave import *
from selenium import webdriver
from POM import loginpom


@given(u'open chrome browser')
def openbrowser(context):
    context.driver = webdriver.Chrome(executable_path="C:\selenium driver\chromedriver_win32\chromedriver.exe")


@when(u'Enter url')
def enterurl(context):
    # context.driver.get("http://www.mitintech.com/webindex/")
    context.driver.get("http://mitintech.com/admin/login/?next=/admin/")
    # raise NotImplementedError(u'STEP: When Enter url')


# @when(u'Enter username "{username}" and password "{password}"')
# def step_impl(context, username, password):
#     context.driver.find_element(By.NAME, "username").send_keys(username)
#     context.driver.find_element(By.NAME, "password").send_keys(password)
#     time.sleep(3)

# @when(u'Enter username and password')
# def step_impl(context):
#     context.driver.find_element(By.NAME,'username').send_keys(loginpom.username)
#     context.driver.find_element(By.NAME,'password').send_keys(loginpom.password)
#     time.sleep(3)

# @when(u'Click on Login button')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()
#
#     time.sleep(3)
#     title = context.driver.title
#     if (title == "Five Finger Admin"):
#         status = ["pass"]
#     else:
#         status = ["fail"]

# @then(u'Login page should be displayed')
# def step_impl(context):
#     title = context.driver.title
#     assert title == "Five Finger Admin"

def cdd(roll):
    username = ''
    password = ''

    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["testing3"]
    rows = sheet.max_row
    # cols = sheet.max_column
    for i in range(1, rows):

        if roll == sheet.cell(row=i + 1, column=2).value:
            username = sheet.cell(row=i + 1, column=3).value
            password = sheet.cell(row=i + 1, column=4).value
            print(username, ':---->', password)
        i = i + 1
    return username, password


@when(u'enter username and password with "{roll}"')
def step_impl(context, roll):
    u, p = cdd(roll)

    context.driver.find_element(By.NAME, 'username').send_keys(u)
    context.driver.find_element(By.NAME, "password").send_keys(p)


# def passfail(status):
#     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#     workbook=openpyxl.load_workbook(path)
#     sheet = workbook.active
#     rows = sheet.max_row
#     cols = sheet.max_column
#     sheet.cell(row=5,column=6).value()
#     workbook.save(path)
#     return status

@when(u'click on login button state"{status}"')
def step_impl(context, status):
    context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()
    time.sleep(3)

    title = context.driver.title
    if title == "Five Fingers Admin | Five Fingers admin site":
        # Admin | Fine Fingers admin site"):
        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        # sheet = workbook.active
        sheet = workbook['testing3']
        cl = sheet["E2"]
        cl.fill = PatternFill("solid", start_color="228B22")
        cl.font = Font(color='ffffff')
        cl.font = Font(size=8)
        cl.font = cl.font.copy(bold=True, italic=True)
        cl.value = "Pass"
        workbook.save(path)

    else:
        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['testing3']
        cl = sheet["E3"]
        cl.value = "Fail"
        cl.fill = PatternFill("solid", start_color="A52A2A")
        cl.font = Font(color='ffffff')
        cl.font = cl.font.copy(bold=True, italic=True)
        cl.font = Font(size=14)

        workbook.save(path)


@then(u'home page should be displayed')
def step_impl(context):
    title = context.driver.title
    if title == "Five Fingers Admin | Five Fingers admin site":
        print(title)
    else:
        get_title = context.driver.title
        # print(get_title)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['testing3']
        sheet['F3']=get_title
        workbook.save(path)
